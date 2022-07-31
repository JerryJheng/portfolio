#===================================================================
# Read "*.ris" file, extract the required data in a excel workbook. 
#===================================================================

from openpyxl import Workbook
import glob
import rispy
import pandas as pd

wb=Workbook()
ws=wb.active
ws.title = 'ref_coding'
#print(wb.sheetnames)

#create cells in memory
for x in range(1,101):
    for y in range(1,101):
        ws.cell(row=x, column=y)

#assign column names
column_names=['Ref_code','Num_EF','Num_Paper','Year','Author','Title','Journal','doi_link']

for i in range(0,len(column_names)):
    #print(column_names[i])
    ws.cell(row=1, column=i+1, value=column_names[i])
#print(ws['A1'].value)

#read .ris file
dirPathPattern = r"C:\Users\*\Desktop\*.ris"
result = glob.glob(dirPathPattern)
row_num=1#column names
for risFile in result:
    with open(risFile,'r',encoding="utf-8") as bibliography_file:
        entries = rispy.load(bibliography_file)
        #print(entries[0])
        for entry in entries:
            #assign row number
            row_num+=1
            try:
                title=entry['primary_title']
            except:
                title=entry['title']
            author=entry['authors']
            au=str(author[0])
            au=au[:au.index(',')]
            try:
                year=entry['year']
            except:
                year=" "
            journal=entry['secondary_title']
            authors_str=""
            for i in range(0,len(author)):
                authors_str=authors_str+author[i]
            #assign ref_code, year, author, title
            au=au + ' et al. '+year
            ref_code_cell=ws.cell(row=row_num,column=1,value=au)
            year_cell=ws.cell(row=row_num,column=4,value=year)
            author_cell=ws.cell(row=row_num,column=5,value=authors_str)
            title_cell=ws.cell(row=row_num,column=6,value=title)
            journal_cell=ws.cell(row=row_num,column=7,value=journal)
            try:
                doi_site='https://www.doi.org/'+ entry['doi']
                if " " in doi_site:
                    doi_temp=str(doi_site)
                    doi_site=doi_temp[:doi_temp.index(" ")]
                #assign ref doi_link
                doi_cell=ws.cell(row=row_num,column=8)
                doi_cell.hyperlink=doi_site
                doi_cell.value=doi_site
                doi_cell.style='Hyperlink'
            except:
                try:
                    url="https://www.webofscience.com/wos/woscc/full-record/"+entry['accession_number']
                    doi_cell=ws.cell(row=row_num,column=8)
                    doi_cell.hyperlink=url
                    doi_cell.value=url
                    doi_cell.style='Hyperlink'
                except:
                    pass

wb.save('test.xlsx')